import pandas as pd
import numpy as np
from faker import Faker
from datetime import datetime, timedelta
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

fake = Faker()
np.random.seed(42)
random.seed(42)

# ── City configs with realistic market characteristics ──────────────────────
CITIES = {
    "New York City": {
        "state": "NY", "avg_price": 185, "price_std": 90, "superhost_rate": 0.22,
        "top_neighborhoods": ["Manhattan", "Brooklyn", "Queens", "Bronx", "Staten Island"],
        "room_mix": {"Entire home/apt": 0.55, "Private room": 0.38, "Shared room": 0.07},
        "avg_stay": 3.2, "market_share": 0.22
    },
    "Los Angeles": {
        "state": "CA", "avg_price": 165, "price_std": 85, "superhost_rate": 0.26,
        "top_neighborhoods": ["Hollywood", "Venice", "Silver Lake", "Santa Monica", "Echo Park"],
        "room_mix": {"Entire home/apt": 0.65, "Private room": 0.30, "Shared room": 0.05},
        "avg_stay": 4.1, "market_share": 0.18
    },
    "Chicago": {
        "state": "IL", "avg_price": 130, "price_std": 65, "superhost_rate": 0.24,
        "top_neighborhoods": ["River North", "Lincoln Park", "Wicker Park", "Logan Square", "South Loop"],
        "room_mix": {"Entire home/apt": 0.58, "Private room": 0.36, "Shared room": 0.06},
        "avg_stay": 3.5, "market_share": 0.14
    },
    "Nashville": {
        "state": "TN", "avg_price": 145, "price_std": 70, "superhost_rate": 0.28,
        "top_neighborhoods": ["Downtown", "East Nashville", "12 South", "Germantown", "The Gulch"],
        "room_mix": {"Entire home/apt": 0.72, "Private room": 0.24, "Shared room": 0.04},
        "avg_stay": 2.8, "market_share": 0.12
    },
    "Miami": {
        "state": "FL", "avg_price": 175, "price_std": 95, "superhost_rate": 0.20,
        "top_neighborhoods": ["South Beach", "Wynwood", "Brickell", "Little Havana", "Coconut Grove"],
        "room_mix": {"Entire home/apt": 0.68, "Private room": 0.27, "Shared room": 0.05},
        "avg_stay": 4.5, "market_share": 0.16
    },
    "Portland": {
        "state": "OR", "avg_price": 110, "price_std": 50, "superhost_rate": 0.30,
        "top_neighborhoods": ["Pearl District", "Alberta Arts", "Division", "Hawthorne", "NW District"],
        "room_mix": {"Entire home/apt": 0.52, "Private room": 0.42, "Shared room": 0.06},
        "avg_stay": 3.8, "market_share": 0.10
    },
    "Austin": {
        "state": "TX", "avg_price": 150, "price_std": 75, "superhost_rate": 0.25,
        "top_neighborhoods": ["Downtown", "East Austin", "South Congress", "Rainey Street", "Hyde Park"],
        "room_mix": {"Entire home/apt": 0.63, "Private room": 0.32, "Shared room": 0.05},
        "avg_stay": 3.0, "market_share": 0.08
    },
}

TRAVEL_PURPOSES = ["Leisure", "Business", "Family vacation", "Remote work", "Events/Festival", "Medical", "Relocation"]
TRAVEL_PURPOSE_WEIGHTS = [0.42, 0.20, 0.16, 0.10, 0.08, 0.02, 0.02]

AGE_GROUPS = ["18-24", "25-34", "35-44", "45-54", "55-64", "65+"]
AGE_WEIGHTS = [0.12, 0.31, 0.24, 0.18, 0.10, 0.05]

GENDERS = ["Male", "Female", "Non-binary", "Prefer not to say"]
GENDER_WEIGHTS = [0.46, 0.48, 0.04, 0.02]

NATIONALITIES = [
    "American", "British", "Canadian", "Australian", "German", "French",
    "Brazilian", "Mexican", "Japanese", "Korean", "Indian", "Chinese"
]
NATIONALITY_WEIGHTS = [0.55, 0.07, 0.06, 0.04, 0.04, 0.03, 0.04, 0.03, 0.03, 0.03, 0.04, 0.04]

PAYMENT_METHODS = ["Credit Card", "Debit Card", "PayPal", "Apple Pay", "Google Pay"]
PAYMENT_WEIGHTS = [0.48, 0.22, 0.16, 0.09, 0.05]

AMENITIES_POOL = [
    "WiFi", "Kitchen", "Air conditioning", "Heating", "Washer", "Dryer",
    "TV", "Free parking", "Pool", "Hot tub", "Gym", "Pet friendly",
    "EV charger", "Workspace", "BBQ grill", "Fireplace"
]

CANCELLATION_POLICIES = ["Flexible", "Moderate", "Strict", "Super Strict"]
CANCEL_WEIGHTS = [0.30, 0.35, 0.28, 0.07]

N_GUESTS = 10000
N_LISTINGS = 2500
N_HOSTS = 1200

# ── 1. GUESTS TABLE ─────────────────────────────────────────────────────────
def generate_guests():
    rows = []
    for i in range(N_GUESTS):
        city_name = random.choices(list(CITIES.keys()), weights=[c["market_share"] for c in CITIES.values()])[0]
        age_group = random.choices(AGE_GROUPS, weights=AGE_WEIGHTS)[0]
        gender = random.choices(GENDERS, weights=GENDER_WEIGHTS)[0]
        nationality = random.choices(NATIONALITIES, weights=NATIONALITY_WEIGHTS)[0]
        join_date = fake.date_between(start_date="-5y", end_date="today")
        verified = random.choices([True, False], weights=[0.78, 0.22])[0]
        total_trips = max(1, int(np.random.negative_binomial(2, 0.3)))
        avg_rating_given = round(random.uniform(3.5, 5.0), 1) if total_trips > 0 else None
        rows.append({
            "guest_id": f"G{i+1:05d}",
            "age_group": age_group,
            "gender": gender,
            "nationality": nationality,
            "primary_destination_city": city_name,
            "join_date": join_date,
            "identity_verified": verified,
            "total_trips_completed": total_trips,
            "average_rating_given": avg_rating_given,
            "preferred_payment_method": random.choices(PAYMENT_METHODS, weights=PAYMENT_WEIGHTS)[0],
            "travel_purpose_most_common": random.choices(TRAVEL_PURPOSES, weights=TRAVEL_PURPOSE_WEIGHTS)[0],
            "superguest_status": total_trips >= 10 and verified,
        })
    return pd.DataFrame(rows)

# ── 2. HOSTS & LISTINGS TABLE ────────────────────────────────────────────────
def generate_listings():
    host_rows = []
    listing_rows = []
    
    for h in range(N_HOSTS):
        city_name = random.choices(list(CITIES.keys()), weights=[c["market_share"] for c in CITIES.values()])[0]
        city = CITIES[city_name]
        is_superhost = random.random() < city["superhost_rate"]
        host_since = fake.date_between(start_date="-8y", end_date="-6m")
        n_listings = random.choices([1, 2, 3, 4, 5], weights=[0.60, 0.20, 0.10, 0.06, 0.04])[0]
        host_id = f"H{h+1:04d}"
        
        host_rows.append({
            "host_id": host_id,
            "host_name": fake.name(),
            "host_city": city_name,
            "host_state": city["state"],
            "superhost": is_superhost,
            "host_since": host_since,
            "total_listings": n_listings,
            "response_rate_pct": round(random.uniform(75, 100) if is_superhost else random.uniform(50, 95), 1),
            "acceptance_rate_pct": round(random.uniform(80, 100) if is_superhost else random.uniform(55, 95), 1),
            "host_rating": round(random.uniform(4.3, 5.0) if is_superhost else random.uniform(3.2, 4.9), 2),
        })
        
        for l in range(n_listings):
            listing_id = f"L{len(listing_rows)+1:05d}"
            room_type = random.choices(list(city["room_mix"].keys()), weights=list(city["room_mix"].values()))[0]
            base_price = max(35, int(np.random.normal(city["avg_price"], city["price_std"])))
            if room_type == "Private room": base_price = int(base_price * 0.55)
            if room_type == "Shared room": base_price = int(base_price * 0.30)
            if is_superhost: base_price = int(base_price * 1.08)
            
            n_amenities = random.randint(5, len(AMENITIES_POOL))
            amenities = random.sample(AMENITIES_POOL, n_amenities)
            
            listing_rows.append({
                "listing_id": listing_id,
                "host_id": host_id,
                "city": city_name,
                "state": city["state"],
                "neighborhood": random.choice(city["top_neighborhoods"]),
                "room_type": room_type,
                "bedrooms": random.choices([0, 1, 2, 3, 4, 5], weights=[0.08, 0.30, 0.30, 0.18, 0.10, 0.04])[0],
                "bathrooms": random.choices([1, 1.5, 2, 2.5, 3], weights=[0.40, 0.20, 0.25, 0.09, 0.06])[0],
                "max_guests": random.choices([1, 2, 3, 4, 5, 6, 8, 10], weights=[0.08, 0.25, 0.15, 0.22, 0.12, 0.10, 0.05, 0.03])[0],
                "nightly_price_usd": base_price,
                "cleaning_fee_usd": random.choices([0, 25, 50, 75, 100, 150], weights=[0.10, 0.15, 0.30, 0.25, 0.15, 0.05])[0],
                "minimum_nights": random.choices([1, 2, 3, 5, 7, 14, 30], weights=[0.20, 0.30, 0.20, 0.12, 0.10, 0.05, 0.03])[0],
                "cancellation_policy": random.choices(CANCELLATION_POLICIES, weights=CANCEL_WEIGHTS)[0],
                "instant_bookable": random.choices([True, False], weights=[0.62, 0.38])[0],
                "listing_rating": round(random.uniform(4.0, 5.0) if is_superhost else random.uniform(3.0, 5.0), 2),
                "total_reviews": max(0, int(np.random.negative_binomial(3, 0.15))),
                "amenities": ", ".join(amenities),
                "listed_date": fake.date_between(start_date=host_since, end_date="today"),
            })
            
            if len(listing_rows) >= N_LISTINGS:
                break
        if len(listing_rows) >= N_LISTINGS:
            break
    
    return pd.DataFrame(host_rows), pd.DataFrame(listing_rows)

# ── 3. BOOKINGS TABLE ────────────────────────────────────────────────────────
def generate_bookings(guests_df, listings_df):
    rows = []
    listing_ids = listings_df["listing_id"].tolist()
    listing_lookup = listings_df.set_index("listing_id")
    guest_ids = guests_df["guest_id"].tolist()
    
    booking_count = 0
    attempt = 0
    while booking_count < N_GUESTS and attempt < N_GUESTS * 3:
        attempt += 1
        guest_id = random.choice(guest_ids)
        listing_id = random.choice(listing_ids)
        listing = listing_lookup.loc[listing_id]
        
        checkin = fake.date_between(start_date="-2y", end_date="+3m")
        city = listing["city"]
        avg_stay = CITIES[city]["avg_stay"]
        nights = int(max(int(listing["minimum_nights"]), int(np.random.negative_binomial(2, 2/avg_stay))))
        checkout = checkin + timedelta(days=nights)
        
        base_total = listing["nightly_price_usd"] * nights
        cleaning = listing["cleaning_fee_usd"]
        airbnb_fee = round(base_total * 0.14, 2)  # ~14% service fee
        total = round(base_total + cleaning + airbnb_fee, 2)
        
        status_choices = ["Completed", "Cancelled by guest", "Cancelled by host", "Upcoming"]
        status_weights = [0.72, 0.14, 0.04, 0.10]
        status = random.choices(status_choices, weights=status_weights)[0]
        
        guest_rating = None
        host_rating_for_guest = None
        if status == "Completed":
            guest_rating = round(random.uniform(2.5, 5.0), 1)
            host_rating_for_guest = round(random.uniform(3.0, 5.0), 1)
        
        travel_purpose = random.choices(TRAVEL_PURPOSES, weights=TRAVEL_PURPOSE_WEIGHTS)[0]
        party_size = random.randint(1, min(listing["max_guests"], 6))
        
        rows.append({
            "booking_id": f"B{booking_count+1:06d}",
            "guest_id": guest_id,
            "listing_id": listing_id,
            "city": city,
            "checkin_date": checkin,
            "checkout_date": checkout,
            "nights_stayed": nights,
            "party_size": party_size,
            "travel_purpose": travel_purpose,
            "booking_status": status,
            "nightly_rate_usd": listing["nightly_price_usd"],
            "cleaning_fee_usd": cleaning,
            "airbnb_service_fee_usd": airbnb_fee,
            "total_paid_usd": total,
            "payment_method": random.choices(PAYMENT_METHODS, weights=PAYMENT_WEIGHTS)[0],
            "booked_days_in_advance": (checkin - fake.date_between(start_date=checkin - timedelta(days=180), end_date=checkin)).days,
            "guest_rating_for_listing": guest_rating,
            "host_rating_for_guest": host_rating_for_guest,
            "booking_source": random.choices(["App", "Web", "API Partner"], weights=[0.58, 0.36, 0.06])[0],
        })
        booking_count += 1
    
    return pd.DataFrame(rows)

# ── 4. SPENDING TABLE ────────────────────────────────────────────────────────
def generate_spending(bookings_df, guests_df):
    completed = bookings_df[bookings_df["booking_status"] == "Completed"].copy()
    guest_lookup = guests_df.set_index("guest_id")
    rows = []
    
    SPEND_CATS = {
        "Dining & Restaurants": (0.85, 20, 250),
        "Groceries & Markets": (0.70, 10, 120),
        "Local Transportation": (0.80, 15, 180),
        "Attractions & Activities": (0.65, 10, 200),
        "Shopping & Retail": (0.55, 10, 300),
        "Nightlife & Entertainment": (0.40, 15, 150),
        "Spa & Wellness": (0.20, 30, 200),
        "Tours & Experiences": (0.35, 25, 250),
    }
    
    for _, booking in completed.iterrows():
        guest = guest_lookup.loc[booking["guest_id"]]
        age_multiplier = {"18-24": 0.75, "25-34": 1.0, "35-44": 1.15, "45-54": 1.25, "55-64": 1.10, "65+": 0.95}
        purpose_multiplier = {"Business": 1.30, "Leisure": 1.0, "Family vacation": 1.20, "Remote work": 0.85,
                               "Events/Festival": 1.35, "Medical": 0.70, "Relocation": 0.60}
        
        age_mult = age_multiplier.get(guest["age_group"], 1.0)
        purpose_mult = purpose_multiplier.get(booking["travel_purpose"], 1.0)
        nights_mult = min(1 + (booking["nights_stayed"] - 1) * 0.15, 2.5)
        party_mult = 1 + (booking["party_size"] - 1) * 0.35
        
        city_cost = {"New York City": 1.35, "Los Angeles": 1.20, "Chicago": 1.05,
                     "Nashville": 1.0, "Miami": 1.15, "Portland": 0.95, "Austin": 1.0}
        city_mult = city_cost.get(booking["city"], 1.0)
        
        spend_id = 0
        for category, (prob, low, high) in SPEND_CATS.items():
            if random.random() < prob:
                base_spend = random.uniform(low * booking["nights_stayed"], high * booking["nights_stayed"])
                adjusted = base_spend * age_mult * purpose_mult * nights_mult * party_mult * city_mult
                adjusted = min(adjusted, high * booking["nights_stayed"] * 3)
                rows.append({
                    "spend_id": f"S{len(rows)+1:07d}",
                    "booking_id": booking["booking_id"],
                    "guest_id": booking["guest_id"],
                    "city": booking["city"],
                    "spend_category": category,
                    "amount_usd": round(adjusted, 2),
                    "nights_stayed": booking["nights_stayed"],
                    "party_size": booking["party_size"],
                    "travel_purpose": booking["travel_purpose"],
                    "age_group": guest["age_group"],
                    "gender": guest["gender"],
                    "nationality": guest["nationality"],
                })
    
    return pd.DataFrame(rows)

# ── Generate all tables ──────────────────────────────────────────────────────
print("Generating guests...")
guests_df = generate_guests()

print("Generating hosts & listings...")
hosts_df, listings_df = generate_listings()

print("Generating bookings...")
bookings_df = generate_bookings(guests_df, listings_df)

print("Generating spending...")
spending_df = generate_spending(bookings_df, guests_df)

print(f"Guests: {len(guests_df):,} | Listings: {len(listings_df):,} | Hosts: {len(hosts_df):,} | Bookings: {len(bookings_df):,} | Spend rows: {len(spending_df):,}")

# ── Write to Excel ───────────────────────────────────────────────────────────
output_path = "/mnt/user-data/outputs/airbnb_synthetic_dataset.xlsx"

HEADER_FILL = PatternFill("solid", start_color="FF5A5F")   # Airbnb red
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="FFF0F0")
NORMAL_FILL = PatternFill("solid", start_color="FFFFFF")
BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
)

def style_sheet(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else NORMAL_FILL
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
    
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)
    
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for df, sheet_name in [
        (guests_df, "Guests"),
        (hosts_df, "Hosts"),
        (listings_df, "Listings"),
        (bookings_df, "Bookings"),
        (spending_df, "Spending"),
    ]:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    wb = writer.book
    for df, sheet_name in [
        (guests_df, "Guests"),
        (hosts_df, "Hosts"),
        (listings_df, "Listings"),
        (bookings_df, "Bookings"),
        (spending_df, "Spending"),
    ]:
        style_sheet(wb[sheet_name], df)

print("Done! File saved.")
